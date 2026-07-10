"""
Cuadre · Shopify → Sage 50
App de Streamlit: sube el export de pedidos de Shopify (.csv o .xlsx) y descarga
el Excel maestro con IVA, comisiones, envíos e intracomunitarias ya calculados.

Despliegue: Streamlit Community Cloud (share.streamlit.io). Requiere requirements.txt.
Restringir acceso: Settings -> Sharing -> privado + invitar emails.
"""
import io
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font

# Países de la UE (ISO-2). Se excluye el país propio (ES) de las intracomunitarias.
EU = {'AT','BE','BG','HR','CY','CZ','DK','EE','FI','FR','DE','GR','HU','IE','IT',
      'LV','LT','LU','MT','NL','PL','PT','RO','SK','SI','SE','ES'}
OWN = 'ES'
EUR = '#,##0.00" €"'

# Cabeceras estándar de un export de pedidos de Shopify (79 columnas, orden fijo).
# Se usan como respaldo cuando el archivo llega SIN fila de cabeceras.
CANON = ['Name', 'Email', 'Financial Status', 'Paid at', 'Fulfillment Status', 'Fulfilled at', 'Accepts Marketing', 'Currency', 'Subtotal', 'Shipping', 'Taxes', 'Total', 'Discount Code', 'Discount Amount', 'Shipping Method', 'Created at', 'Lineitem quantity', 'Lineitem name', 'Lineitem price', 'Lineitem compare at price', 'Lineitem sku', 'Lineitem requires shipping', 'Lineitem taxable', 'Lineitem fulfillment status', 'Billing Name', 'Billing Street', 'Billing Address1', 'Billing Address2', 'Billing Company', 'Billing City', 'Billing Zip', 'Billing Province', 'Billing Country', 'Billing Phone', 'Shipping Name', 'Shipping Street', 'Shipping Address1', 'Shipping Address2', 'Shipping Company', 'Shipping City', 'Shipping Zip', 'Shipping Province', 'Shipping Country', 'Shipping Phone', 'Notes', 'Note Attributes', 'Cancelled at', 'Payment Method', 'Payment Reference', 'Refunded Amount', 'Vendor', 'Outstanding Balance', 'Employee', 'Location', 'Device ID', 'Id', 'Tags', 'Risk Level', 'Source', 'Lineitem discount', 'Tax 1 Name', 'Tax 1 Value', 'Tax 2 Name', 'Tax 2 Value', 'Tax 3 Name', 'Tax 3 Value', 'Tax 4 Name', 'Tax 4 Value', 'Tax 5 Name', 'Tax 5 Value', 'Phone', 'Receipt Number', 'Duties', 'Billing Province Name', 'Shipping Province Name', 'Payment ID', 'Payment Terms Name', 'Next Payment Due At', 'Payment References']

# Campos que necesita el cálculo (por si hay que resolverlos por posición)
KEY_FIELDS = ['Financial Status', 'Subtotal', 'Total']


def numf(v):
    """Convierte a número tolerando texto y formatos EU/US."""
    if v is None:
        return 0.0
    s = str(v).strip()
    if s == '':
        return 0.0
    s = ''.join(ch for ch in s if ch in '0123456789.,-')
    if ',' in s and '.' in s:
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')   # 1.234,56 (EU)
        else:
            s = s.replace(',', '')                     # 1,234.56 (US)
    elif ',' in s:
        s = s.replace(',', '.')                         # 134,00
    try:
        return float(s)
    except ValueError:
        return 0.0


def quarter(created):
    """Devuelve (trimestre '2ºT/26', fecha 'dd/mm/yyyy') desde 'Created at' ISO."""
    s = str(created or '')
    if len(s) < 7:
        return ('', '')
    y, mo, da = s[0:4], s[5:7], s[8:10]
    try:
        m = int(mo)
    except ValueError:
        return ('', '')
    q = (m - 1) // 3 + 1
    return (f"{q}ºT/{y[2:]}", f"{da}/{mo}/{y}")


def build_workbook(df, iva_pct, com_pct, fija):
    """Procesa el DataFrame y devuelve (bytes_xlsx, totales)."""
    IVA = iva_pct / 100.0
    COM = com_pct / 100.0
    div = 1 + IVA
    cols = list(df.columns)

    # --- Respaldo: si el archivo llegó SIN cabeceras (faltan los nombres clave)
    #     pero tiene la estructura estándar de Shopify, se aplican los nombres
    #     canónicos por posición. Así funciona con o sin fila de cabeceras. ---
    restored = False
    low = {str(c).strip().lower(): c for c in cols}

    def by_name(field):
        if field in cols:
            return field
        return low.get(field.lower())

    if not all(by_name(k) for k in KEY_FIELDS):
        # El archivo no trae los nombres clave -> respaldo por posición (orden Shopify)
        if len(cols) >= 33:
            cols = [CANON[i] if i < len(CANON) else str(cols[i]) for i in range(len(cols))]
            df = df.copy()
            df.columns = cols
            low = {str(c).strip().lower(): c for c in cols}
            restored = True
        if not all((k in cols) for k in KEY_FIELDS):
            raise ValueError(
                "No reconozco este archivo como un export de pedidos de Shopify "
                "(no encuentro columnas como 'Financial Status', 'Subtotal' o 'Total', "
                "ni por nombre ni por posición). Exporta los pedidos desde Shopify "
                "(Pedidos → Exportar) e inténtalo de nuevo."
            )

    # Mapa campo canónico -> columna real del archivo (exacta o insensible a mayúsculas)
    FIELDS = ['Name', 'Financial Status', 'Subtotal', 'Shipping', 'Total', 'Created at',
              'Billing Country', 'Lineitem quantity', 'Lineitem name',
              'Lineitem price', 'Refunded Amount']
    COLMAP = {f: by_name(f) for f in FIELDS}

    def g(row, field):
        c = COLMAP.get(field)
        return '' if c is None else row[c]

    calc_headers = ['PEDIDO', 'TRIMESTRE', 'FECHA', 'BASE PROD (700)', 'IVA PROD (477.21)',
                    'TOTAL COBRADO (430)', 'BASE ENVÍO', 'IVA ENVÍO',
                    'COMISIÓN SHOPIFY', 'NETO RECIBIDO', 'INTRACOMUNITARIA', 'PAÍS UE']
    pedidos = [calc_headers + cols]

    byQ, byC, byP = {}, {}, {}
    T = dict(n=0, baseP=0.0, ivaP=0.0, baseE=0.0, ivaE=0.0, com=0.0, neto=0.0,
             ref=0.0, intraBase=0.0, intraTot=0.0)

    for _, r in df.iterrows():
        status = str(g(r, 'Financial Status') or '').strip().lower()
        country = str(g(r, 'Billing Country') or '').strip().upper()
        pedido = str(g(r, 'Name') or '').strip()   # nº de pedido (#1013)
        calc = [pedido] + [''] * 11

        if status != '' and status != 'refunded':
            sub = numf(g(r, 'Subtotal'))
            tot = numf(g(r, 'Total'))
            baseP = sub / div
            ivaP = sub - baseP
            net = max(0.0, tot - sub)
            baseE = net / div
            ivaE = net - baseE
            com = tot * COM + fija
            neto = tot - com
            qt, dt = quarter(g(r, 'Created at'))
            intra = bool(country) and country != OWN and country in EU
            calc = [pedido, qt, dt, baseP, ivaP, tot, baseE, ivaE, com, neto,
                    ('SÍ' if intra else 'NO'), (country if intra else '')]
            T['n'] += 1
            T['baseP'] += baseP; T['ivaP'] += ivaP
            T['baseE'] += baseE; T['ivaE'] += ivaE
            T['com'] += com; T['neto'] += neto
            if qt:
                s = byQ.setdefault(qt, dict(baseP=0.0, ivaP=0.0, baseE=0.0, ivaE=0.0, n=0))
                s['baseP'] += baseP; s['ivaP'] += ivaP
                s['baseE'] += baseE; s['ivaE'] += ivaE; s['n'] += 1
            if intra:
                c = byC.setdefault(country, dict(base=0.0, iva=0.0, tot=0.0, n=0))
                c['base'] += baseP + baseE; c['iva'] += ivaP + ivaE
                c['tot'] += tot; c['n'] += 1
                T['intraBase'] += baseP + baseE; T['intraTot'] += tot

        ref = numf(g(r, 'Refunded Amount'))
        if ref > 0:
            T['ref'] += ref

        if status != 'refunded':
            name = str(g(r, 'Lineitem name') or '').strip()
            if name:
                qty = numf(g(r, 'Lineitem quantity'))
                price = numf(g(r, 'Lineitem price'))
                p = byP.setdefault(name, dict(qty=0.0, rev=0.0))
                p['qty'] += qty; p['rev'] += qty * price

        raw = ['' if (pd.isna(r[c]) if not isinstance(r[c], str) else False) else r[c] for c in cols]
        pedidos.append(calc + raw)

    # ---- hojas resumen ----
    if T['n'] == 0:
        raise ValueError(
            "El archivo se leyó, pero no encontré ningún pedido válido "
            "(revisa que la columna de estado tenga 'paid' y que haya importes). "
            "¿Seguro que es el export de pedidos de Shopify?"
        )
    resumen = [['Trimestre', 'Base productos', 'IVA productos', 'Base envíos',
                'IVA envíos', 'Base total', 'IVA a ingresar', 'Nº pedidos']]
    for k in sorted(byQ):
        s = byQ[k]
        resumen.append([k, s['baseP'], s['ivaP'], s['baseE'], s['ivaE'],
                        s['baseP'] + s['baseE'], s['ivaP'] + s['ivaE'], s['n']])
    resumen.append(['TOTAL', T['baseP'], T['ivaP'], T['baseE'], T['ivaE'],
                    T['baseP'] + T['baseE'], T['ivaP'] + T['ivaE'], T['n']])

    intra = [['País', 'Nº pedidos', 'Base imponible', 'IVA', 'Total']]
    for c in sorted(byC):
        x = byC[c]
        intra.append([c, x['n'], x['base'], x['iva'], x['tot']])
    intra.append(['TOTAL UE (no-ES)', sum(x['n'] for x in byC.values()),
                  T['intraBase'], sum(x['iva'] for x in byC.values()), T['intraTot']])

    stock = [['Producto', 'Uds vendidas', 'Ingresos brutos', 'Precio medio',
              'Coste unitario', 'Beneficio bruto']]
    sQty = sRev = 0.0
    for n in sorted(byP):
        p = byP[n]
        sQty += p['qty']; sRev += p['rev']
        avg = (p['rev'] / p['qty']) if p['qty'] else 0.0
        stock.append([n, p['qty'], p['rev'], avg, '', ''])  # coste/beneficio -> fórmula
    stock.append(['TOTAL', sQty, sRev, '', '', ''])

    iva_total = T['ivaP'] + T['ivaE']
    prod_gross = T['baseP'] + T['ivaP']
    env_gross = T['baseE'] + T['ivaE']
    fact = prod_gross + env_gross
    benef = fact - iva_total - T['com']
    pyl = [
        ['CUENTA DE RESULTADOS', ''],
        ['Ventas producto (IVA incl.)', prod_gross],
        ['Envíos cobrados (IVA incl.)', env_gross],
        ['Total facturado', fact],
        ['IVA repercutido (a Hacienda)', iva_total],
        ['Comisiones Shopify', T['com']],
        ['Devoluciones (informativo)', T['ref']],
        ['Beneficio bruto (antes de gastos operativos)', benef],
        ['', ''],
        ['Añade tus gastos, campañas y costes de producto para el beneficio neto.', ''],
    ]

    params = [
        ['PARÁMETROS', 'Valor', 'Nota'],
        ['IVA', IVA, '21% general'],
        ['Comisión Shopify', COM, 'sobre el Total'],
        ['Comisión fija (€)', fija, 'por transacción'],
        ['País propio', OWN, 'excluido de intracomunitarias'],
        ['Cuenta base producto', '700', ''],
        ['Cuenta IVA repercutido', '477.21', ''],
        ['Cuenta clientes', '430', 'Total cobrado'],
        ['Cuenta comisiones', '626', ''],
    ]

    # ---- construir XLSX ----
    wb = Workbook()
    wb.remove(wb.active)

    def add(name, aoa, money=(), pct=()):
        ws = wb.create_sheet(name)
        for ri, row in enumerate(aoa, 1):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if ri > 1 and isinstance(val, (int, float)) and not isinstance(val, bool):
                    if (ci - 1) in money:
                        cell.number_format = EUR
                    if (ci - 1) in pct:
                        cell.number_format = '0.00%'
        return ws

    add('PARAMETROS', params, pct=(1,))
    wp = add('PEDIDOS', pedidos, money=(3, 4, 5, 6, 7, 8, 9))
    for ci in range(1, len(pedidos[0]) + 1):
        wp.column_dimensions[wp.cell(row=1, column=ci).column_letter].width = 14 if ci <= 12 else 16
    add('RESUMEN_IVA_TRIMESTRAL', resumen, money=(1, 2, 3, 4, 5, 6))
    add('INTRACOMUNITARIAS', intra, money=(2, 3, 4))

    ws = add('STOCK_PRODUCTOS', stock, money=(2, 3, 4, 5))
    nS = len(stock)  # header + productos + TOTAL
    for R in range(2, nS):  # filas de producto
        ws[f'F{R}'] = f'=IF(E{R}="","",C{R}-B{R}*E{R})'
        ws[f'F{R}'].number_format = EUR
    ws[f'F{nS}'] = f'=SUM(F2:F{nS-1})'
    ws[f'F{nS}'].number_format = EUR

    add('PYL_BENEFICIOS', pyl, money=(1,))

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    totals = dict(n=T['n'], iva=iva_total, com=T['com'], neto=T['neto'],
                  fact=fact, benef=benef, intra=T['intraTot'], ref=T['ref'],
                  prod=prod_gross, env=env_gross, hasIntra=len(byC), restored=restored)
    return bio.getvalue(), totals


def read_orders(filename, raw_bytes):
    """Lee un export de Shopify (.csv o .xlsx) de forma robusta:
    detecta codificación y delimitador del CSV, y elimina filas vacías."""
    import io as _io
    name = str(filename).lower()
    if name.endswith('.csv') or name.endswith('.txt'):
        text = None
        for enc in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
            try:
                text = raw_bytes.decode(enc)
                break
            except Exception:
                continue
        if text is None:
            text = raw_bytes.decode('utf-8', errors='replace')
        import csv as _csv
        # Detecta el delimitador probando cuál genera más columnas en la 1ª fila
        sep, best = ',', 0
        for cand in (',', ';', '\t', '|'):
            try:
                first = next(_csv.reader(_io.StringIO(text), delimiter=cand))
                if len(first) > best:
                    best, sep = len(first), cand
            except Exception:
                continue
        df = pd.read_csv(_io.StringIO(text), dtype=str, keep_default_na=False, sep=sep)
    else:
        df = pd.read_excel(_io.BytesIO(raw_bytes), dtype=str).fillna('')
    df.columns = [str(c).strip() for c in df.columns]
    df = df[~df.apply(lambda row: all(str(x).strip() == '' for x in row), axis=1)]
    return df.reset_index(drop=True)


# =========================== UI de Streamlit ===========================
st.set_page_config(page_title="Cuadre · Shopify → Sage 50", page_icon="📊", layout="centered")

st.title("Cuadre")
st.caption("Shopify → Sage 50 · IVA, comisiones, envíos e intracomunitarias automáticos")

up = st.file_uploader("Sube el export de pedidos de Shopify (.csv o .xlsx)",
                      type=['csv', 'xlsx', 'xls'])

c1, c2, c3 = st.columns(3)
iva = c1.number_input("IVA (%)", value=21.0, step=0.1, format="%.2f")
com = c2.number_input("Comisión Shopify (%)", value=2.1, step=0.01, format="%.2f")
fija = c3.number_input("Comisión fija (€)", value=0.30, step=0.01, format="%.2f")
st.caption("La comisión se aplica sobre el Total. El envío se computa por lo realmente "
           "cobrado (Total − Subtotal); si fue gratis no genera IVA. Los pedidos "
           "`refunded` se excluyen de los cálculos.")

if up is not None:
    try:
        df = read_orders(up.name, up.getvalue())

        data, T = build_workbook(df, iva, com, fija)
        if T.get('restored'):
            st.info("El archivo venía sin fila de cabeceras; he aplicado el formato "
                    "estándar de Shopify por posición de columna.")
        st.success(f"Procesados {T['n']} pedidos.")

        m1, m2, m3 = st.columns(3)
        m1.metric("IVA a ingresar (Modelo 303)", f"{T['iva']:,.2f} €")
        m2.metric("Total facturado", f"{T['fact']:,.2f} €")
        m3.metric("Beneficio bruto", f"{T['benef']:,.2f} €")
        m4, m5, m6 = st.columns(3)
        m4.metric("Ventas producto", f"{T['prod']:,.2f} €")
        m5.metric("Comisiones Shopify", f"{T['com']:,.2f} €")
        m6.metric("Ventas intracomunitarias", f"{T['intra']:,.2f} €")

        if T['hasIntra']:
            st.info("Detectadas ventas intracomunitarias: revisa el régimen OSS (B2C) "
                    "o la exención por NIF-VIES (B2B) en la hoja INTRACOMUNITARIAS.")

        st.download_button(
            "⬇️ Descargar SHOPIFY_SAGE50_MAESTRO.xlsx",
            data=data,
            file_name="SHOPIFY_SAGE50_MAESTRO.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        st.error(f"No pude procesar el archivo: {e}")

st.divider()
st.caption("Los datos se procesan en el servidor de la app para generar el Excel y no se "
           "almacenan. Acceso restringido a los emails autorizados.")
